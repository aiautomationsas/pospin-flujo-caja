"""Excel importer — adapted from the CLI prototype to store data in Supabase."""
from datetime import datetime, date
from pathlib import Path

import openpyxl


def safe_num(val) -> float:
    """Convert cell value to float, returning 0 if None or invalid."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _find_sheet(wb, keyword: str):
    """Find a sheet whose name contains the keyword (case-insensitive)."""
    for name in wb.sheetnames:
        if keyword.upper() in name.upper():
            return wb[name]
    return None


def parse_excel(file_path: str) -> dict:
    """Parse an Excel file with up to 3 sheets.
    
    Returns dict with keys: bancos, flujo, compromisos, sheets_found.
    Each value is a list of dicts or None if sheet not found.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {"sheets_found": wb.sheetnames, "bancos": None, "flujo": None, "compromisos": None}

    # --- BANCOS SEMANAL ---
    ws_bancos = _find_sheet(wb, "BANCOS")
    if ws_bancos:
        result["bancos"] = _parse_bancos(ws_bancos)

    # --- FLUJO DE CAJA PROYECTADO ---
    ws_flujo = _find_sheet(wb, "FLUJO")
    if ws_flujo:
        result["flujo"] = _parse_flujo(ws_flujo)

    # --- COMPROMISOS ESPECIALES ---
    ws_comp = _find_sheet(wb, "COMPROMISOS")
    if ws_comp:
        result["compromisos"] = _parse_compromisos(ws_comp)

    return result


def _parse_bancos(ws) -> list[dict]:
    """Parse BANCOS SEMANAL sheet. Rows 3-54, columns A-D for accounts."""
    rows = []
    for row_num in range(3, 55):
        semana = ws.cell(row=row_num, column=1).value
        if semana is None:
            continue
        try:
            semana_int = int(semana)
        except (ValueError, TypeError):
            continue

        cuentas = []
        for col in range(2, 5):  # columns B, C, D
            val = safe_num(ws.cell(row=row_num, column=col).value)
            cuentas.append(val)

        rows.append({
            "semana": semana_int,
            "saldos": cuentas,  # list of saldo values, one per account column
        })
    return rows


def _parse_flujo(ws) -> dict:
    """Parse FLUJO DE CAJA PROYECTADO sheet.
    
    Flexible parsing: detects week columns by scanning header row.
    Extracts facturas (rows 11-22), egresos (rows 27-48).
    """
    # Detect week columns: scan row 4 for numeric week numbers
    week_cols = []
    for col in range(5, 30):  # scan up to column AC
        val = ws.cell(row=4, column=col).value
        if val is not None:
            try:
                week_cols.append((col, int(val)))
            except (ValueError, TypeError):
                pass

    if not week_cols:
        return {"facturas": [], "egresos": [], "semanas": []}

    semanas = [w for _, w in week_cols]

    # Facturas (rows 11-22)
    facturas = []
    for row_num in range(11, 23):
        cliente = ws.cell(row=row_num, column=1).value
        factura = ws.cell(row=row_num, column=2).value
        fecha_venc = ws.cell(row=row_num, column=3).value
        valor = safe_num(ws.cell(row=row_num, column=4).value)

        if not cliente and not factura:
            continue

        # Recaudo por semana
        recaudo_semanal = {}
        for col, semana_num in week_cols:
            val = safe_num(ws.cell(row=row_num, column=col).value)
            if val > 0:
                recaudo_semanal[semana_num] = val

        facturas.append({
            "cliente": str(cliente).strip() if cliente else "",
            "factura": str(factura).strip() if factura else "",
            "fecha_venc": _to_date(fecha_venc),
            "valor": valor,
            "recaudo_semanal": recaudo_semanal,
        })

    # Egresos (rows 27-48)
    egresos = []
    for row_num in range(27, 49):
        cat = ws.cell(row=row_num, column=1).value
        if not cat:
            continue
        valores = {}
        for col, semana_num in week_cols:
            val = safe_num(ws.cell(row=row_num, column=col).value)
            if val > 0:
                valores[semana_num] = val
        if valores:
            egresos.append({
                "categoria": str(cat).strip(),
                "valores": valores,
            })

    return {"facturas": facturas, "egresos": egresos, "semanas": semanas}


def _parse_compromisos(ws) -> list[dict]:
    """Parse COMPROMISOS ESPECIALES sheet. Rows 3-14, columns B-E."""
    rows = []
    for row_num in range(3, 15):
        tercero = ws.cell(row=row_num, column=2).value
        if not tercero:
            continue
        fecha = ws.cell(row=row_num, column=3).value
        valor = safe_num(ws.cell(row=row_num, column=4).value)
        prioridad = ws.cell(row=row_num, column=5).value

        rows.append({
            "tercero": str(tercero).strip(),
            "fecha": _to_date(fecha),
            "valor": valor,
            "prioridad": _normalize_prioridad(str(prioridad).strip() if prioridad else "media"),
        })
    return rows


def _to_date(val) -> str | None:
    """Convert a cell value to ISO date string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return str(val)


def _normalize_prioridad(val: str) -> str:
    """Normalize prioridad to allowed values."""
    v = val.lower()
    if v in ("alta", "media", "baja"):
        return v
    return "media"


def import_to_db(client, parsed_data: dict) -> int:
    """Insert parsed Excel data into Supabase tables. Returns count of records inserted."""
    count = 0

    # --- Import cuentas (from bancos sheet) ---
    # We don't know account names from Excel alone, so skip auto-creating cuentas.
    # User must configure accounts in the app first.

    # --- Import facturas y recaudos ---
    flujo = parsed_data.get("flujo")
    if flujo:
        count += _import_facturas_y_recaudos(client, flujo)

    # --- Import egresos ---
    if flujo:
        count += _import_egresos(client, flujo)

    # --- Import compromisos ---
    compromisos = parsed_data.get("compromisos")
    if compromisos:
        count += _import_compromisos(client, compromisos)

    return count


def _import_facturas_y_recaudos(client, flujo: dict) -> int:
    """Import facturas and their recaudos from parsed flujo data."""
    count = 0
    facturas_data = flujo.get("facturas", [])

    for fac in facturas_data:
        if not fac["cliente"]:
            continue

        # Upsert cliente
        cliente_resp = client.table("clientes").select("id").eq("nombre", fac["cliente"]).execute()
        if cliente_resp.data:
            cliente_id = cliente_resp.data[0]["id"]
        else:
            ins = client.table("clientes").insert({"nombre": fac["cliente"]}).execute()
            cliente_id = ins.data[0]["id"]

        # Upsert factura
        factura_resp = client.table("facturas").select("id").eq("numero", fac["factura"]).eq("cliente_id", cliente_id).execute()
        if factura_resp.data:
            factura_id = factura_resp.data[0]["id"]
        else:
            fecha_emision = fac.get("fecha_venc") or date.today().isoformat()
            ins = client.table("facturas").insert({
                "cliente_id": cliente_id,
                "numero": fac["factura"],
                "fecha_emision": fecha_emision,
                "fecha_vencimiento": fac.get("fecha_venc") or date.today().isoformat(),
                "valor": fac["valor"],
                "estado": "pendiente",
            }).execute()
            factura_id = ins.data[0]["id"]

        # Insert recaudos por semana
        for semana_num, valor in fac.get("recaudo_semanal", {}).items():
            # Find semana_id
            sem_resp = client.table("semanas").select("id").eq("numero", semana_num).limit(1).execute()
            if sem_resp.data:
                semana_id = sem_resp.data[0]["id"]
                # Check if recaudo already exists
                existing = client.table("recaudos").select("id").eq(
                    "semana_id", semana_id
                ).eq("factura_id", factura_id).execute()
                if not existing.data:
                    client.table("recaudos").insert({
                        "semana_id": semana_id,
                        "factura_id": factura_id,
                        "valor": valor,
                        "fecha": date.today().isoformat(),
                    }).execute()
                    count += 1
    return count


def _import_egresos(client, flujo: dict) -> int:
    """Import egresos from parsed flujo data."""
    count = 0
    egresos_data = flujo.get("egresos", [])

    for eg in egresos_data:
        # Upsert categoria
        cat_resp = client.table("categorias_egreso").select("id").eq("nombre", eg["categoria"]).execute()
        if cat_resp.data:
            categoria_id = cat_resp.data[0]["id"]
        else:
            ins = client.table("categorias_egreso").insert({
                "nombre": eg["categoria"],
                "tipo": "terceros",
            }).execute()
            categoria_id = ins.data[0]["id"]

        # Insert egresos por semana
        for semana_num, valor in eg.get("valores", {}).items():
            sem_resp = client.table("semanas").select("id").eq("numero", semana_num).limit(1).execute()
            if sem_resp.data:
                semana_id = sem_resp.data[0]["id"]
                # Upsert (semana_id, categoria_id is unique)
                existing = client.table("egresos").select("id").eq(
                    "semana_id", semana_id
                ).eq("categoria_id", categoria_id).execute()
                if existing.data:
                    client.table("egresos").update({"valor": valor}).eq("id", existing.data[0]["id"]).execute()
                else:
                    client.table("egresos").insert({
                        "semana_id": semana_id,
                        "categoria_id": categoria_id,
                        "valor": valor,
                    }).execute()
                count += 1
    return count


def _import_compromisos(client, compromisos: list[dict]) -> int:
    """Import compromisos from parsed data."""
    count = 0
    for comp in compromisos:
        if not comp.get("fecha"):
            continue
        # Check if already exists (by tercero + fecha)
        existing = client.table("compromisos").select("id").eq(
            "tercero", comp["tercero"]
        ).eq("fecha", comp["fecha"]).execute()
        if not existing.data:
            client.table("compromisos").insert({
                "tercero": comp["tercero"],
                "fecha": comp["fecha"],
                "valor": comp["valor"],
                "prioridad": comp["prioridad"],
                "estado": "pendiente",
            }).execute()
            count += 1
    return count


def register_import(client, filename: str, sheets: str, count: int, success: bool):
    """Log an import to the importaciones table."""
    client.table("importaciones").insert({
        "archivo": filename,
        "hojas": sheets,
        "registros": count,
        "exitosa": success,
    }).execute()
