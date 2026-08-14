"""Export functions for Excel and PDF reports."""
import io
from datetime import datetime


def export_excel(
    proyeccion_data: list[dict],
    saldos_cuenta: list[dict],
    recaudo_pendiente: list[dict],
    obligaciones_pendientes: list[dict] = None,
) -> io.BytesIO:
    """Generate an Excel workbook with projection data.

    Args:
        proyeccion_data: list of dicts with keys: semana, anio, saldo_inicial, recaudo, egresos, obligaciones, saldo_final
        saldos_cuenta: list of dicts with keys: nombre, banco, numero, saldo
        recaudo_pendiente: list of dicts with keys: cliente, facturas (list of {numero, valor, pendiente}), total_pendiente
        obligaciones_pendientes: list of dicts representing pending obligations

    Returns: BytesIO with .xlsx content.
    """
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1: Proyección Semanal
    ws1 = wb.active
    ws1.title = "Proyección Semanal"
    ws1.append([
        "Semana",
        "Saldo Inicial",
        "Recaudos Proyectados",
        "Egresos Recurrentes",
        "Obligaciones (CxP)",
        "Saldo Final Proyectado",
        "Estado Liquidez",
    ])
    for p in proyeccion_data:
        egres_rec = p.get("egresos_recurrente", 0) + p.get("egresos_real", 0)
        oblig = p.get("obligaciones", 0)
        saldo_fin = p.get("saldo_final", 0)
        ws1.append([
            f"Semana {p.get('semana', '')} ({p.get('anio', '')})",
            p.get("saldo_inicial", 0),
            p.get("recaudo", 0),
            egres_rec,
            oblig,
            saldo_fin,
            "Déficit" if saldo_fin < 0 else "Superávit",
        ])

    # Sheet 2: Saldos por Cuenta
    ws2 = wb.create_sheet("Saldos por Cuenta")
    ws2.append(["Cuenta", "Banco", "Número", "Saldo"])
    for sc in saldos_cuenta:
        ws2.append([
            sc.get("nombre", ""),
            sc.get("banco", ""),
            sc.get("numero", ""),
            sc.get("saldo", 0),
        ])

    # Sheet 3: Recaudo Pendiente
    ws3 = wb.create_sheet("Recaudo Pendiente")
    ws3.append(["Cliente", "Factura", "Valor", "Pendiente"])
    for r in recaudo_pendiente:
        cliente = r.get("cliente", "")
        for fac in r.get("facturas", []):
            ws3.append([
                cliente,
                fac.get("numero", ""),
                fac.get("valor", 0),
                fac.get("pendiente", 0),
            ])

    # Sheet 4: Obligaciones Pendientes
    if obligaciones_pendientes:
        ws4 = wb.create_sheet("Obligaciones Pendientes")
        ws4.append([
            "Tercero",
            "Concepto",
            "Monto Total",
            "Saldo Pendiente",
            "Fecha Vencimiento",
            "Fecha Programada",
            "Prioridad",
            "Estado",
        ])
        for ob in obligaciones_pendientes:
            ws4.append([
                ob.get("tercero", ""),
                ob.get("concepto", ""),
                ob.get("monto_total", 0),
                ob.get("saldo_pendiente", 0),
                str(ob.get("fecha_vencimiento") or ""),
                str(ob.get("fecha_programada_pago") or ""),
                ob.get("prioridad", ""),
                ob.get("estado", ""),
            ])

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_pdf(
    proyeccion_data: list[dict],
    saldos_cuenta: list[dict],
    recaudo_pendiente: list[dict],
    obligaciones_pendientes: list[dict] = None,
) -> io.BytesIO:
    """Generate a PDF report with projection data.

    Args: same as export_excel.
    Returns: BytesIO with .pdf content.
    """
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Flujo de Caja Proyectado - Grupo Pospin", ln=True, align="C")

    # Date
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    pdf.ln(5)

    # Projection table
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Proyección Semanal", ln=True)
    pdf.set_font("Helvetica", "", 8)

    # Table header
    pdf.set_fill_color(220, 220, 220)
    headers = ["Semana", "Saldo Inicial", "Recaudos", "Egresos Rec.", "Obligaciones", "Saldo Final"]
    col_widths = [30, 30, 30, 30, 30, 30]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    # Table rows
    deficits = []
    for p in proyeccion_data:
        saldo_final = p.get("saldo_final", 0)
        egres_rec = p.get("egresos_recurrente", 0) + p.get("egresos_real", 0)
        oblig = p.get("obligaciones", 0)
        row = [
            f"Sem {p.get('semana', '')} ({p.get('anio', '')})",
            _fmt_num(p.get("saldo_inicial", 0)),
            _fmt_num(p.get("recaudo", 0)),
            _fmt_num(egres_rec),
            _fmt_num(oblig),
            _fmt_num(saldo_final),
        ]
        if saldo_final < 0:
            pdf.set_text_color(255, 0, 0)
            deficits.append(p)
        for i, val in enumerate(row):
            pdf.cell(col_widths[i], 6, val, border=1, align="R" if i > 0 else "L")
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    pdf.ln(5)

    # Account balances
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Saldos por Cuenta", ln=True)
    pdf.set_font("Helvetica", "", 9)

    headers2 = ["Cuenta", "Banco", "Número", "Saldo"]
    col_widths2 = [50, 40, 40, 45]
    pdf.set_fill_color(220, 220, 220)
    for i, h in enumerate(headers2):
        pdf.cell(col_widths2[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    for sc in saldos_cuenta:
        row = [
            str(sc.get("nombre", "")),
            str(sc.get("banco", "")),
            str(sc.get("numero", "")),
            _fmt_num(sc.get("saldo", 0)),
        ]
        for i, val in enumerate(row):
            pdf.cell(col_widths2[i], 6, val, border=1, align="R" if i == 3 else "L")
        pdf.ln()

    # Obligations section
    if obligaciones_pendientes:
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Obligaciones Pendientes (Cuentas por Pagar)", ln=True)
        pdf.set_font("Helvetica", "", 8)

        headers3 = ["Tercero", "Concepto", "Saldo Pendiente", "Fecha Prog.", "Prioridad", "Estado"]
        col_widths3 = [40, 50, 32, 23, 18, 17]
        pdf.set_fill_color(220, 220, 220)
        for i, h in enumerate(headers3):
            pdf.cell(col_widths3[i], 7, h, border=1, fill=True, align="C")
        pdf.ln()

        for ob in obligaciones_pendientes:
            tercero = str(ob.get("tercero", ""))[:22]
            concepto = str(ob.get("concepto", ""))[:28]
            saldo = _fmt_num(ob.get("saldo_pendiente", 0))
            f_prog = str(ob.get("fecha_programada_pago") or "")[:10]
            prio = str(ob.get("prioridad", ""))[:8]
            est = str(ob.get("estado", ""))[:8]

            row = [tercero, concepto, saldo, f_prog, prio, est]
            for i, val in enumerate(row):
                pdf.cell(col_widths3[i], 6, val, border=1, align="R" if i == 2 else "L")
            pdf.ln()

    # Alerts section
    if deficits:
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(255, 0, 0)
        pdf.cell(0, 8, "ALERTAS - Semanas con Déficit", ln=True)
        pdf.set_font("Helvetica", "", 9)
        for d in deficits:
            pdf.cell(0, 6,
                f"Semana {d.get('semana', '')} ({d.get('anio', '')}): Déficit de {_fmt_num(abs(d.get('saldo_final', 0)))}",
                ln=True)
        pdf.set_text_color(0, 0, 0)

    # Output
    output = io.BytesIO()
    output.write(pdf.output())
    output.seek(0)
    return output


def _fmt_num(val) -> str:
    """Format number for PDF display."""
    try:
        num = int(round(float(val)))
        return f"${num:,}".replace(",", ".")
    except (ValueError, TypeError):
        return "$0"

